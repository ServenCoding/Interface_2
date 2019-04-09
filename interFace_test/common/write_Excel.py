# -*- coding: utf-8 -*-
# @Time    : 2019/1/25  20:41
# @Author  : 陆平！！
# @FileName: write_Excel.py
# @Software: PyCharm

import openpyxl

def copy_Excel(old_Excel,new_Excel):
    wb2 = openpyxl.Workbook()
    wb2.save(new_Excel)
    #读取数据
    wb1 = openpyxl.load_workbook(old_Excel)
    wb2 = openpyxl.load_workbook(new_Excel)
    sheets1 = wb1.sheetnames
    sheets2 = wb2.sheetnames
    sheet1 = wb1[sheets1[0]]
    sheet2 = wb2[sheets2[0]]
    max_row = sheet1.max_row
    max_column = sheet1.max_column

    for m in list(range(1,max_row+1)):
        for n in list(range(97, 97 + max_column)):  # chr(97)='a'
            n = chr(n)           # ASCII字符
            i = '%s%d' % (n, m)  # 单元格编号
            cell1 = sheet1[i].value  # 获取data单元格数据
            sheet2[i].value = cell1  # 赋值到test单元格
        wb2.save(new_Excel)  # 保存数据
    wb1.close()        # 关闭excel
    wb2.close()
    # '''
    # 设置宽高
    # '''
    # col_width = 256 * 20
    # try:
    #     for i in itertools.count():
    #         sheet2.col(i).width = col_width
    # except ValueError:
    #     pass
    # default_book_style = workbook.default_style
    # default_book_style.font.height = 20 * 36

class Unit():

    def __init__(self, filename):
        self.filename = filename
        self.wb = openpyxl.load_workbook(self.filename)
        self.ws = self.wb.active      #激活Sheets

    def write_Sheets(self,row_num,col_num,value):
        '''
        写入Excel
        '''
        self.ws.cell(row_num,col_num).value = value
        self.wb.save(self.filename)
if __name__ == "__main__":
    copy_Excel('D:\project_address_1\interface4\The_test_case\interfaceTest.xlsx','D:\project_address_1\interface4\The_test_case\q2.xlsx')
    wf = Unit('D:\project_address_1\interface4\The_test_case\q2.xlsx')
    wf.write_Sheets(3,2,"sdsds")